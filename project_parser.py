"""
Pipe-level parser: reads per-pipe repair data from each project sheet in the
daily Excel workbook (every sheet other than the summary/report sheets).

Each pipe is recorded as a fixed-size block (an anchor cell containing a
"Repair ... Rate :" label, with the pipe number, repair amount, repair
count, and surface state a few rows below it). A block's ratio value sits
either under the "Coating" side or the "Clutch" side of the block — whichever
one is filled in determines the pipe's repair category; empty blocks are
unused template slots and are skipped.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from validators import coerce_number, normalize_spaces

# Known non-project sheets in the workbook (summary/report/chart tabs).
NON_PROJECT_SHEETS = {
    "Daily Activity w Contents",
    "Daily Activity w Contents (2)",
    "Daily Repair Rate",
    "Daily Repair Rate (Old)",
    "Daily Repair Rate (2)",
    "Daily Chart",
    "Sheet1",
}

_LABEL_PREFIX = "repairrate"
# Label formats seen across sheets: "129' 8\" ft" (feet+inches) and
# "134.54 FT" / "134.54 ft" / "65 feet" (decimal feet, "ft" or spelled out).
_LENGTH_FT_IN_RE = re.compile(r"(\d+)\s*'\s*(\d+)?\s*\"")
_LENGTH_DECIMAL_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(?:ft|feet)\b", re.IGNORECASE)

PIPE_NO_ROW_OFFSET = 4
DATA_ROW_OFFSET = 5
# A newer, more compact block template (seen first on the "06-131" pier/
# stage-style sheets) uses the same columns but a shorter block: pipe
# no./length two rows below the anchor, amount/count/state three rows
# below, instead of four/five. Tried only as a fallback (see _parse_sheet)
# so sheets that already match the standard offsets are unaffected.
PIPE_NO_ROW_OFFSET_COMPACT = 2
DATA_ROW_OFFSET_COMPACT = 3
# A third block variant (seen on "10-108 (60x.375)") has the ratio value
# three columns right of the anchor instead of directly under it
# ("Coating", offset 0) or two over ("Clutch", offset 2) — every block on
# that sheet is labeled "Coating" in its own text, so this offset is
# treated as another Coating position. pipe_no/amount stay at the standard
# row offsets there, so no row-offset change is needed for it.
RATIO_COL_OFFSET_ALT = 3


@dataclass
class ProjectParseReport:
    parsed_rows: int = 0
    parsed_sheets: int = 0
    skipped_blocks: int = 0
    warnings: list[str] = field(default_factory=list)


def _is_label_cell(value: object) -> bool:
    if not isinstance(value, str):
        return False
    return value.replace(" ", "").lower().startswith(_LABEL_PREFIX)


def _parse_length_ft(value: object) -> float | None:
    """Parse a pipe-length label into a decimal feet value.

    Handles both "129' 8\" ft" (feet+inches) and "134.54 FT" (decimal feet)
    label formats — different project sheets use different templates.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    match = _LENGTH_FT_IN_RE.search(text)
    if match:
        feet = int(match.group(1))
        inches = int(match.group(2)) if match.group(2) else 0
        return round(feet + inches / 12, 4)

    match = _LENGTH_DECIMAL_RE.search(text)
    if match:
        return round(float(match.group(1)), 4)

    return None


def _load_sheet_grid(ws) -> dict[tuple[int, int], object]:
    """Materialize non-empty cells in a single fast pass (values_only avoids
    the per-call overhead of random cell access in read-only worksheets)."""
    grid: dict[tuple[int, int], object] = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for col_idx, value in enumerate(row, start=1):
            if value is not None:
                grid[(row_idx, col_idx)] = value
    return grid


def _extract_block(
    grid: dict[tuple[int, int], object], anchor_row: int, anchor_col: int, pipe_no_offset: int, data_offset: int
) -> tuple[float, float | None, float, float | None, str | None] | None:
    """Try to pull pipe_no/length/amount/count/surface_state out of a block
    using the given row offsets. Returns None if pipe_no or repair_amount
    is missing at these offsets (either a genuinely unused template slot,
    or the wrong offset pair for this sheet's template)."""
    pipe_no_row = anchor_row + pipe_no_offset
    data_row = anchor_row + data_offset

    pipe_no = coerce_number(grid.get((pipe_no_row, anchor_col + 1)))
    repair_amount = coerce_number(grid.get((data_row, anchor_col + 1)))
    if pipe_no is None or repair_amount is None:
        return None

    pipe_length_ft = _parse_length_ft(grid.get((pipe_no_row, anchor_col)))
    repair_count = coerce_number(grid.get((data_row, anchor_col + 3)))
    surface_state = normalize_spaces(grid.get((data_row, anchor_col + 4)))
    return pipe_no, pipe_length_ft, repair_amount, repair_count, surface_state


def _parse_sheet(ws, sheet_name: str, report_date_iso: str, report: ProjectParseReport) -> list[dict]:
    grid = _load_sheet_grid(ws)
    anchors = [pos for pos, value in grid.items() if _is_label_cell(value)]

    rows: list[dict] = []
    for anchor_row, anchor_col in anchors:
        coating_ratio = coerce_number(grid.get((anchor_row, anchor_col)))
        clutch_ratio = coerce_number(grid.get((anchor_row, anchor_col + 2)))
        alt_ratio = coerce_number(grid.get((anchor_row, anchor_col + RATIO_COL_OFFSET_ALT)))

        if coating_ratio is not None:
            repair_ratio, repair_category = coating_ratio, "Coating"
        elif clutch_ratio is not None:
            repair_ratio, repair_category = clutch_ratio, "Clutch"
        elif alt_ratio is not None:
            repair_ratio, repair_category = alt_ratio, "Coating"
        else:
            continue  # unused template slot, not an actual pipe entry

        extracted = _extract_block(grid, anchor_row, anchor_col, PIPE_NO_ROW_OFFSET, DATA_ROW_OFFSET)
        if extracted is None:
            extracted = _extract_block(
                grid, anchor_row, anchor_col, PIPE_NO_ROW_OFFSET_COMPACT, DATA_ROW_OFFSET_COMPACT
            )
        if extracted is None:
            report.skipped_blocks += 1
            continue

        pipe_no, pipe_length_ft, repair_amount, repair_count, surface_state = extracted

        rows.append(
            {
                "date": report_date_iso,
                "project_sheet": sheet_name,
                "block_cell": f"{get_column_letter(anchor_col)}{anchor_row}",
                "pipe_no": int(pipe_no),
                "pipe_length_ft": pipe_length_ft,
                "repair_amount": repair_amount,
                "repair_ratio": repair_ratio,
                "repair_count": int(repair_count) if repair_count is not None else None,
                "repair_category": repair_category,
                "surface_state": surface_state or None,
            }
        )
    return rows


def parse_project_pipe_repairs(
    file: str | Path | BinaryIO, report_date: object
) -> tuple[pd.DataFrame, ProjectParseReport]:
    report = ProjectParseReport()
    report_date_iso = pd.to_datetime(report_date).date().isoformat()

    wb = load_workbook(file, read_only=True, data_only=True)

    all_rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in NON_PROJECT_SHEETS:
            continue
        ws = wb[sheet_name]
        sheet_rows = _parse_sheet(ws, sheet_name, report_date_iso, report)
        if sheet_rows:
            all_rows.extend(sheet_rows)
            report.parsed_sheets += 1

    report.parsed_rows = len(all_rows)
    if not all_rows:
        report.warnings.append("No pipe-level data found in any project sheet.")

    return pd.DataFrame(all_rows), report
