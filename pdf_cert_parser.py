"""
Module 2: parses SDI (Steel Dynamics) Metallurgical Certification PDFs and
shapes the extracted coil data into the same "Heat / Coil Test Report"
layout already used for the Excel-supplier's data (see
`module 2 exam files/79488 Standard Check.xlsx`, sheet "Report") — same
column order, same section headers, same heat-grouped "Row Count" rollup.

Field extraction uses two techniques depending on how reliably pdfplumber
reconstructs each part of the page:
  - Coil #, Heat #, Material Spec, Length, Width, Gauge, Weight (A) come
    from `extract_tables()` — SDI lays these out in real table cells, and
    the two-column labels ("Ship To" / "Sold To" boxes) that surround them
    on the page never get mixed into these particular tables.
  - Yield/Tensile/Elongation and the Charpy (CVN) readings sit in a
    two-column "Sample 1 / Sample 2" layout that pdfplumber does NOT
    recover as a clean table (it interleaves the two columns line by
    line). Those are pulled via regex directly off the page text instead,
    which stays reliable because each value still trails its own label on
    one line even though the *next* line belongs to the other column.

Only one real SDI cert has been seen so far, so this is built to fail
loudly (missing field -> reported error) rather than guess, so a
different SDI template doesn't silently produce wrong numbers.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Order matches "79488 Standard Check.xlsx" -> sheet "Report" exactly.
# Section spans: Part Information A:C, Elements D:U, Mechanical Properties V:AB.
REPORT_COLUMNS = [
    "Heat No",
    "Coil / FULL_TAG_NUM",
    "Row Count",
    "C",
    "Mn",
    "P",
    "S",
    "Si",
    "Ni",
    "Cr",
    "Mo",
    "Cu",
    "Al",
    "V",
    "Nb",
    "Ti",
    "N",
    "Pb",
    "B",
    "Sn",
    "soln_Cert Stipulatn Code",
    "Yield KSI",
    "Tensile",
    "Elongation%",
    "CVN Ind1",
    "CVN Ind2",
    "CVN Ind3",
    "CVN Avg",
]

# The PDF's Ladle Chemical Analysis table also has "Ca", which the target
# report doesn't include — left out on purpose.
CHEMISTRY_ELEMENTS = ["C", "Mn", "P", "S", "Si", "Ni", "Cr", "Mo", "Cu", "Al", "V", "Nb", "Ti", "N", "Pb", "B", "Sn"]

_MECHANICAL_TO_COLUMN = {
    "yield_ksi": "Yield KSI",
    "tensile_ksi": "Tensile",
    "elongation_pct": "Elongation%",
    "cvn_ind1": "CVN Ind1",
    "cvn_ind2": "CVN Ind2",
    "cvn_ind3": "CVN Ind3",
    "cvn_avg": "CVN Avg",
}


def _normalize_label(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", text.lower())


def _table_label_values(tables: list[list[list[str | None]]]) -> dict[str, str]:
    """Flatten every 'label, value, label, value, ...' table row into a
    lookup dict, keyed by normalized label. None cells (column-alignment
    gaps SDI's table layout leaves between pairs) are dropped before
    pairing, not paired-through, so a gap doesn't shift later pairs."""
    pairs: dict[str, str] = {}
    for table in tables:
        for row in table:
            cells = [c for c in row if c is not None]
            for i in range(0, len(cells) - 1, 2):
                label, value = cells[i], cells[i + 1]
                if isinstance(label, str) and label.strip():
                    pairs[_normalize_label(label)] = str(value).strip()
    return pairs


def _extract_number(text: str | None) -> float | None:
    if not text:
        return None
    match = re.search(r"[\d,]+\.?\d*", text)
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def _find_chemistry_table(tables: list[list[list[str | None]]]) -> dict[str, float | None]:
    for table in tables:
        if not table or not table[0]:
            continue
        header = table[0]
        if sum(1 for h in header if h in CHEMISTRY_ELEMENTS) >= len(CHEMISTRY_ELEMENTS) - 2:
            values = table[1] if len(table) > 1 else []
            return {h.strip(): _extract_number(v) for h, v in zip(header, values) if h}
    return {}


@dataclass
class PdfCertResult:
    source_file: str
    fields: dict[str, object] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


def parse_sdi_cert_pdf(file: str | Path | BinaryIO, source_file: str) -> PdfCertResult:
    """Parse one SDI Metallurgical Certification PDF (page 1 for the main
    fields, page 2 only for the Charpy average, matching this cert's
    2-page layout)."""
    result = PdfCertResult(source_file=source_file)

    try:
        with pdfplumber.open(file) as pdf:
            page1 = pdf.pages[0]
            page1_text = page1.extract_text() or ""
            tables = page1.extract_tables()
            page2_text = pdf.pages[1].extract_text() if len(pdf.pages) > 1 else ""
    except Exception as exc:
        result.errors.append(f"Could not read PDF: {exc}")
        return result

    labels = _table_label_values(tables)

    heat_no = labels.get("heat")
    coil_number = labels.get("coil")
    if not heat_no:
        result.errors.append("Heat # not found")
    if not coil_number:
        result.errors.append("Coil # not found")

    result.fields["heat_no"] = heat_no
    result.fields["coil_number"] = coil_number
    result.fields["material_spec"] = labels.get("materialspec") or None
    result.fields["length_ft"] = _extract_number(labels.get("length"))
    result.fields["width_in"] = _extract_number(labels.get("width"))
    result.fields["gauge_in"] = _extract_number(labels.get("gauge"))
    result.fields["weight_lb"] = _extract_number(labels.get("weighta"))

    chemistry = _find_chemistry_table(tables)
    if not chemistry:
        result.errors.append("Ladle Chemical Analysis table not found")
    for element in CHEMISTRY_ELEMENTS:
        result.fields[element] = chemistry.get(element)

    yield_match = re.search(r"Yield\s*Strength\s+([\d.]+)\s*KSI", page1_text, re.IGNORECASE)
    tensile_match = re.search(r"Tensile\s*Strength\s+([\d.]+)\s*KSI", page1_text, re.IGNORECASE)
    elong_match = re.search(r"Total\s*Elongation\s+([\d.]+)\s*%", page1_text, re.IGNORECASE)
    result.fields["yield_ksi"] = float(yield_match.group(1)) if yield_match else None
    result.fields["tensile_ksi"] = float(tensile_match.group(1)) if tensile_match else None
    result.fields["elongation_pct"] = float(elong_match.group(1)) if elong_match else None
    if not yield_match:
        result.errors.append("Yield Strength not found")
    if not tensile_match:
        result.errors.append("Tensile Strength not found")
    if not elong_match:
        result.errors.append("Total Elongation not found")

    # Charpy (CVN), best-effort: the three "#1/#2/#3" energy readings are
    # the only "FT-LB" values on page 1; the average is the only one on
    # page 2 in this cert's layout.
    cvn_values = [float(m) for m in re.findall(r"(\d+(?:\.\d+)?)\s*FT-LB", page1_text)]
    cvn_avg_values = [float(m) for m in re.findall(r"(\d+(?:\.\d+)?)\s*FT-LB", page2_text)]
    result.fields["cvn_ind1"] = cvn_values[0] if len(cvn_values) > 0 else None
    result.fields["cvn_ind2"] = cvn_values[1] if len(cvn_values) > 1 else None
    result.fields["cvn_ind3"] = cvn_values[2] if len(cvn_values) > 2 else None
    result.fields["cvn_avg"] = cvn_avg_values[0] if cvn_avg_values else None

    return result


def build_report_rows(results: list[PdfCertResult]) -> pd.DataFrame:
    """Group parsed certs by Heat No (same coil-per-heat rollup as the
    Excel-supplier's Standard Check report) and shape into REPORT_COLUMNS."""
    valid = [r for r in results if r.fields.get("heat_no") and r.fields.get("coil_number")]
    groups: dict[str, list[PdfCertResult]] = {}
    for r in valid:
        groups.setdefault(str(r.fields["heat_no"]), []).append(r)

    rows = []
    for heat_no, group in groups.items():
        coil_numbers = list(dict.fromkeys(str(r.fields["coil_number"]) for r in group))
        row: dict[str, object] = {
            "Heat No": heat_no,
            "Coil / FULL_TAG_NUM": " | ".join(coil_numbers),
            "Row Count": len(group),
            "soln_Cert Stipulatn Code": next(
                (r.fields.get("material_spec") for r in group if r.fields.get("material_spec")), None
            ),
        }
        for element in CHEMISTRY_ELEMENTS:
            values = [r.fields[element] for r in group if r.fields.get(element) is not None]
            row[element] = sum(values) / len(values) if values else None
        for internal_key, column in _MECHANICAL_TO_COLUMN.items():
            values = [r.fields[internal_key] for r in group if r.fields.get(internal_key) is not None]
            row[column] = sum(values) / len(values) if values else None
        rows.append(row)

    if not rows:
        return pd.DataFrame(columns=REPORT_COLUMNS)
    return pd.DataFrame(rows, columns=REPORT_COLUMNS).sort_values("Heat No").reset_index(drop=True)


def build_report_workbook(df: pd.DataFrame) -> bytes:
    """Write REPORT_COLUMNS data into the same section-header + bold-column
    layout as the Standard Check report's "Report" sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    bold = Font(bold=True)
    last_col = get_column_letter(len(REPORT_COLUMNS))

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "PDF Cert Report"
    ws["A1"].font = Font(bold=True, size=14)

    ws.merge_cells("A3:C3")
    ws["A3"] = "Part Information"
    ws.merge_cells("D3:T3")
    ws["D3"] = "Elements"
    ws.merge_cells("V3:AB3")
    ws["V3"] = "Mechanical Properties"
    for coord in ("A3", "D3", "V3"):
        ws[coord].font = bold

    for col_idx, col_name in enumerate(REPORT_COLUMNS, start=1):
        ws.cell(row=4, column=col_idx, value=col_name).font = bold

    for row_offset, row in enumerate(df.itertuples(index=False), start=5):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_offset, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
