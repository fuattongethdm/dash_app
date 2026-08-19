from __future__ import annotations

import re
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string

from validators import (
    ValidationReport,
    coerce_number,
    coerce_report_date,
    normalize_spaces,
    normalize_status,
    validate_dataframe,
)


SHEET_CANDIDATES = ["Daily Repair Rate", "Daily Repair Rate (2)", "Daily Repair Rate (Old)"]

REPORT_DATE_CELL = "P1"

OLD_FORMAT_SECTIONS = [
    {"production_type": "Coil", "start_row": 4, "end_row": 25},
]

NEW_FORMAT_SECTIONS = [
    {
        "production_type": "Plate",
        "title_cell": "A3",
        "title_text": "Ongoing & Recently Completed Productions with Plate",
        "start_row": 5,
        "end_row": 14,
    },
    {
        "production_type": "Coil",
        "title_cell": "A15",
        "title_text": "Ongoing & Recently Completed Productions with Coil",
        "start_row": 17,
        "end_row": 41,
    },
]

COLUMN_MAP = {
    "project_no": "B",
    "dimensions": "E",
    "qty": "H",
    "project_total_pipe_length": "I",
    "repaired_pipes_total_length": "J",
    "repaired_spiral_length": "K",
    "total_repair_amount": "L",
    "total_repair_amount_incl_skelp": "M",
    "project_status": "N",
    "repair_ratio": "O",
    "repair_ratio_incl_skelp": "P",
}

# Below the daily table, the same sheet keeps a running "Overall Repair
# Rates" archive of projects that have been completed and removed from the
# active list above. Excel's own "Overall Repair Rate (2026)" summary cell
# (Daily Repair Rate!H44, formula =C61) sums this 2026 row range together
# with the active daily table — without it, the app's overall ratio only
# reflects currently-active projects and skews away from Excel's figure.
ARCHIVE_COLUMN_MAP = {
    "project_no": "B",
    "repaired_spiral_length": "F",
    "total_repair_amount": "G",
    "total_repair_amount_incl_skelp": "H",
}

ARCHIVE_CURRENT_YEAR_SECTION = {"start_row": 154, "end_row": 217}


def _select_sheet(wb):
    available = [name for name in SHEET_CANDIDATES if name in wb.sheetnames]
    if not available:
        return None

    dated = []
    for name in available:
        report_date = coerce_report_date(wb[name][REPORT_DATE_CELL].value)
        dated.append((report_date, name))

    dated.sort(key=lambda item: (item[0] is not None, item[0] or date.min), reverse=True)
    return dated[0][1]


def _is_new_format(ws) -> bool:
    for section in NEW_FORMAT_SECTIONS:
        title = normalize_spaces(ws[section["title_cell"]].value).lower()
        expected = normalize_spaces(section["title_text"]).lower()
        if title != expected:
            return False
    return True


# Some rows have the dimension typed into the project number by mistake,
# e.g. "12-112 (54 x 1.2)" where the dimensions column already separately
# holds "Ø54"x1.2"". Strip that redundant suffix so project_no stays a
# stable identifier — legitimate text qualifiers like "(PT)"/"(HDD)" don't
# match this (numeric-only) pattern and are left alone.
_BAKED_IN_DIMENSION_RE = re.compile(r"\s*\(\s*\d+(?:\.\d+)?\s*x\s*\d+(?:\.\d+)?\s*\)\s*$", re.IGNORECASE)


def _clean_project_no(project_no: str) -> str:
    return _BAKED_IN_DIMENSION_RE.sub("", project_no).strip()


def _production_type(default_type: str, project_no: str) -> str:
    lowered = project_no.lower()
    if "(pt)" in lowered or "plate" in lowered:
        return "Plate"
    return default_type


def parse_daily_repair_rate(wb: Workbook) -> tuple[pd.DataFrame, ValidationReport]:
    """``wb`` is an already-opened openpyxl Workbook (``load_workbook(file,
    read_only=True, data_only=True)``) — the caller opens it once and shares
    it with the other two parsers instead of each re-opening the same file
    from scratch (see pages/home.py:handle_upload)."""
    report = ValidationReport()

    sheet_name = _select_sheet(wb)
    if sheet_name is None:
        report.add_check("Sheet found?", False)
        report.add_error(f"Daily Repair Rate sheet not found. Expected sheet names: {', '.join(SHEET_CANDIDATES)}")
        return pd.DataFrame(), report

    report.add_check("Sheet found?", True)
    ws = wb[sheet_name]

    report_date = coerce_report_date(ws[REPORT_DATE_CELL].value)
    report.add_check("Date found?", report_date is not None)
    if report_date is None:
        report.add_error(f"Report date is empty or invalid: {sheet_name}!{REPORT_DATE_CELL}")
        return pd.DataFrame(), report

    sections = NEW_FORMAT_SECTIONS if _is_new_format(ws) else OLD_FORMAT_SECTIONS

    # Single sequential pass instead of per-cell bracket access — random
    # cell access on a read_only worksheet re-scans from the start of the
    # sheet on every call, which is what made the pipe-level parser take
    # 73s before it was switched to this same pattern (see project_parser.py
    # and parse_repair_rate_archive below, which already used it here).
    col_index = {key: column_index_from_string(letter) for key, letter in COLUMN_MAP.items()}
    min_row = min(section["start_row"] for section in sections)
    max_row = max(section["end_row"] for section in sections)
    grid: dict[tuple[int, int], object] = {}
    rows_iter = ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True)
    for row_idx, row_values in enumerate(rows_iter, start=min_row):
        for col_idx, value in enumerate(row_values, start=1):
            if value is not None:
                grid[(row_idx, col_idx)] = value

    rows: list[dict[str, object]] = []
    for section in sections:
        for excel_row in range(section["start_row"], section["end_row"] + 1):
            project_no = _clean_project_no(normalize_spaces(grid.get((excel_row, col_index["project_no"]))))
            dimensions = normalize_spaces(grid.get((excel_row, col_index["dimensions"])))

            if not project_no or not dimensions:
                continue

            row = {
                "date": report_date.isoformat(),
                "production_type": _production_type(section["production_type"], project_no),
                "project_no": project_no,
                "dimensions": dimensions,
                "project_status": normalize_status(grid.get((excel_row, col_index["project_status"]))),
                "excel_row": excel_row,
            }

            for col_name in [
                "qty",
                "project_total_pipe_length",
                "repaired_pipes_total_length",
                "repaired_spiral_length",
                "total_repair_amount",
                "total_repair_amount_incl_skelp",
                "repair_ratio",
                "repair_ratio_incl_skelp",
            ]:
                row[col_name] = coerce_number(grid.get((excel_row, col_index[col_name])))

            rows.append(row)

    df = pd.DataFrame(rows)
    if not df.empty:
        ordered = [
            "date",
            "production_type",
            "project_no",
            "dimensions",
            "qty",
            "project_total_pipe_length",
            "repaired_pipes_total_length",
            "repaired_spiral_length",
            "total_repair_amount",
            "total_repair_amount_incl_skelp",
            "project_status",
            "repair_ratio",
            "repair_ratio_incl_skelp",
            "excel_row",
        ]
        df = df[ordered]

    return df, validate_dataframe(df, report)


def parse_repair_rate_archive(wb: Workbook) -> pd.DataFrame:
    """Parse the current-year "Overall Repair Rates" archive table beneath
    the daily table (completed projects removed from the active list).

    ``wb`` is an already-opened Workbook, shared with the other parsers —
    see parse_daily_repair_rate's docstring.

    Returns a frame shaped for database.upsert_historical_baselines, or an
    empty frame if the sheet/section isn't found — callers should treat
    this as best-effort supplementary data, not block the main import on it.
    """
    sheet_name = _select_sheet(wb)
    if sheet_name is None:
        return pd.DataFrame()
    ws = wb[sheet_name]

    section = ARCHIVE_CURRENT_YEAR_SECTION
    col_index = {key: column_index_from_string(letter) for key, letter in ARCHIVE_COLUMN_MAP.items()}

    # Single sequential pass instead of per-cell bracket access — random
    # cell access on a read_only worksheet re-scans from the start of the
    # sheet on every call, which is what made the pipe-level parser take
    # 73s before it was switched to this same pattern (see project_parser.py).
    grid: dict[tuple[int, int], object] = {}
    rows_iter = ws.iter_rows(min_row=section["start_row"], max_row=section["end_row"], values_only=True)
    for row_idx, row in enumerate(rows_iter, start=section["start_row"]):
        for col_idx, value in enumerate(row, start=1):
            if value is not None:
                grid[(row_idx, col_idx)] = value

    rows: list[dict[str, object]] = []
    for excel_row in range(section["start_row"], section["end_row"] + 1):
        project_no = normalize_spaces(grid.get((excel_row, col_index["project_no"])))
        if not project_no:
            continue

        spiral_length = coerce_number(grid.get((excel_row, col_index["repaired_spiral_length"])))
        repair_amount = coerce_number(grid.get((excel_row, col_index["total_repair_amount"])))
        repair_amount_incl_skelp = coerce_number(grid.get((excel_row, col_index["total_repair_amount_incl_skelp"])))
        if spiral_length is None or repair_amount is None:
            continue

        rows.append(
            {
                "project_no": project_no,
                "dimensions": "Archived",
                "repaired_spiral_length": spiral_length,
                "total_repair_amount": repair_amount,
                "total_repair_amount_incl_skelp": (
                    repair_amount_incl_skelp if repair_amount_incl_skelp is not None else repair_amount
                ),
                "project_status": "Completed",
            }
        )

    return pd.DataFrame(rows)
