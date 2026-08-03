"""
Module 2: parses the Excel-supplier's (JSW) coil cert export — a wide,
one-row-per-coil sheet (see `module 2 exam files/JD Fields Cert
6-10-26.xlsx`, sheet "Cert") — into the same PdfCertResult shape the SDI
PDF parser produces, so both sources feed the same build_report_rows/
build_report_workbook pipeline in pdf_cert_parser.py and end up in one
identical "Report" layout.

The sheet holding this data isn't always named "Cert" — JSW's exports
vary. Every sheet in the workbook is checked for the required column
headers (order and position don't matter, just presence), and the first
one that has them all is used, rather than assuming a fixed sheet name.

Unlike the PDF, this is already a clean table — column lookup by header
name, no text-position guessing needed.
"""

from __future__ import annotations

from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from pdf_cert_parser import CHEMISTRY_ELEMENTS, PdfCertResult

# Internal field name -> source column header in the "Cert" sheet.
_COLUMN_MAP = {
    "heat_no": "HEAT_NUMBER",
    "coil_number": "FULL_TAG_NUM",
    "material_spec": "soln_Cert Stipulatn Code",
    "yield_ksi": "Yield KSI",
    "tensile_ksi": "Tensile",
    "elongation_pct": "Elongation%",
    "cvn_ind1": "CVN Ind1",
    "cvn_ind2": "CVN Ind2",
    "cvn_ind3": "CVN Ind3",
    "cvn_avg": "CVN Avg",
}
for _element in CHEMISTRY_ELEMENTS:
    _COLUMN_MAP[_element] = _element


def _to_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _find_cert_sheet(wb) -> tuple[str, dict[str, int]] | None:
    """Return (sheet_name, header->column_index) for the first sheet whose
    header row has every column _COLUMN_MAP needs, or None if none do."""
    best_sheet: str | None = None
    best_missing: list[str] | None = None
    for sheet_name in wb.sheetnames:
        header = next(wb[sheet_name].iter_rows(values_only=True), None)
        if header is None:
            continue
        col_index = {name: idx for idx, name in enumerate(header) if name}
        missing = [col for col in _COLUMN_MAP.values() if col not in col_index]
        if not missing:
            return sheet_name, col_index
        if best_missing is None or len(missing) < len(best_missing):
            best_sheet, best_missing = sheet_name, missing
    if best_sheet is not None:
        raise ValueError(
            f"No sheet has all the expected coil-cert columns. Closest match: "
            f"'{best_sheet}' is missing {', '.join(best_missing)}."
        )
    return None


def parse_jsw_cert_excel(file: str | Path | BinaryIO, source_file: str) -> list[PdfCertResult]:
    """One PdfCertResult per data row (= per coil), from whichever sheet in
    the workbook has the expected coil-cert columns (sheet name itself is
    not assumed — see module docstring)."""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        result = PdfCertResult(source_file=source_file)
        result.errors.append(f"Could not open Excel file: {exc}")
        return [result]

    try:
        match = _find_cert_sheet(wb)
    except ValueError as exc:
        result = PdfCertResult(source_file=source_file)
        result.errors.append(str(exc))
        return [result]

    if match is None:
        result = PdfCertResult(source_file=source_file)
        result.errors.append(f"No usable sheet found (checked: {', '.join(wb.sheetnames)})")
        return [result]

    sheet_name, col_index = match
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # header row, already consumed by _find_cert_sheet's own read

    results: list[PdfCertResult] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        coil_number = row[col_index["FULL_TAG_NUM"]]
        if coil_number in (None, ""):
            continue  # blank trailing row

        row_source = f"{source_file} [{sheet_name}] (row {row_idx})"
        result = PdfCertResult(source_file=row_source)

        heat_no = row[col_index["HEAT_NUMBER"]]
        if not heat_no:
            result.errors.append("HEAT_NUMBER not found")
        result.fields["heat_no"] = str(heat_no) if heat_no is not None else None
        result.fields["coil_number"] = str(coil_number)
        result.fields["material_spec"] = row[col_index["soln_Cert Stipulatn Code"]]

        for internal_key, column in _COLUMN_MAP.items():
            if internal_key in ("heat_no", "coil_number", "material_spec"):
                continue
            result.fields[internal_key] = _to_number(row[col_index[column]])

        results.append(result)

    if not results:
        empty = PdfCertResult(source_file=source_file)
        empty.errors.append("No coil rows found")
        return [empty]
    return results
